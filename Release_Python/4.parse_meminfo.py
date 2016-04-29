#!/usr/bin/python
import argparse
import errno
import itertools
import os
import re
import shutil
import subprocess
from subprocess import check_output
from time import strftime

from openpyxl import Workbook
from openpyxl.cell import get_column_letter

# from openpyxl.utils import get_column_letter
from openpyxl.compat import range
from openpyxl.styles import Color, PatternFill, Font, Alignment, Border, Side

CMD_PACKAGES = "adb shell pm list packages -f"
CMD_MEMINFO = "adb shell dumpsys meminfo"
# used for generating dir's name
CMD_DEVICE = "adb shell getprop ro.product.device"
CMD_VERSION = "adb shell getprop ro.build.version.incremental"

SYSTEM_PROC_WHITE_LIST = ['system', 'android.process.media', 'android.process.acore']
GOOGLE_PROC_WHITE_LIST = ['com.google', 'com.android.facelock', 'com.android.vending', 'com.android.chrome']

QCOM_TITLE = ['Memory', 'Total RAM', 'Free RAM', 'Kernel', 'Native',
         'Module', 'GMS', 'Qcom', 'Third', 'System apps',
         'Packages', 'installed', 'gms', 'qcom', 'third', 'system']

MTK_TITLE = ['Memory', 'Total RAM', 'Free RAM', 'Kernel', 'Native',
             'Module', 'GMS', 'MTK', 'Third', 'System apps',
             'Packages', 'installed', 'gms', 'mtk', 'third', 'system']

PLATFORM_PKGS = []
platform_flag = 0

meminfo_file = 'meminfo.txt'
packages_file = 'packages.txt'
out_parse_table = 'parse_table.txt'
out_parse_xlsx = 'parse_diff_info.xlsx'

original_dir = ""
diff_dir = ""


def main():
    global original_dir, diff_dir, out_parse_table
    global platform_flag, PLATFORM_PKGS

    parser = argparse.ArgumentParser(formatter_class=argparse.RawDescriptionHelpFormatter, epilog='''
Example of use:
  platform: 0(Qualcomm)/1(MTK)
  comparison mode: $ python parse_meminfo.py [-p platform:0/1] -f original_dir -d diff_dir
  analysis mode:   $ python parse_meminfo.py [-p platform:0/1] [-o outdir]
''')
    parser.add_argument('-p', '--platform', nargs='?', default='0', help="0: Qualcomm; 1: MTK")
    parser.add_argument('-f', '--file', nargs='*', help="the original directory's name")
    parser.add_argument('-d', '--diff', nargs='*', help="the diff directory's name")
    parser.add_argument('-o', '--outdir', nargs='?',  # default='parse_result.txt',
                        help='For analysis mode, save result in the custom directory')
    args = parser.parse_args()
    # 0: qualcomm; 1: MTK
    platform_flag = args.platform
    if int(platform_flag):
        PLATFORM_PKGS = ["com.mediatek", "com.mtk"]
    else:
        PLATFORM_PKGS = ["com.qti", "com.qualcomm"]

    if args.file is None:
        # prepare output directory's name
        try:
            if args.outdir is None:
                device_name = check_output(CMD_DEVICE, shell=True)
                version_name = check_output(CMD_VERSION, shell=True)
                cur_time = strftime("%m%d%H%M")
                outdir = "{}_{}_{}".format(device_name.strip(), version_name.strip(), cur_time.strip())
            else:
                outdir = str(args.outdir)

            # Run command with arguments and return its output as a byte string.
            file_meminfo = check_output(CMD_MEMINFO, shell=True)
            write_to_file("{}/{}".format(outdir, meminfo_file), file_meminfo)
            system_mem, filtered_procs_mem = get_file_meminfo(file_meminfo)

            installed_pkgs_str = check_output(CMD_PACKAGES, shell=True)
            write_to_file("{}/{}".format(outdir, packages_file), installed_pkgs_str)
            pkgs_dic = get_packages_dic(installed_pkgs_str)
        except subprocess.CalledProcessError:
            print "*** Error: please check your device state.\n"
            # Programatically stop execution of python script
            raise SystemError(0)

        groups_mem_kb, groups_mem = get_procs_attr_group(filtered_procs_mem, pkgs_dic)
        out_parse_table = "{}/{}".format(outdir, out_parse_table)
        print_mem_table(out_parse_table, system_mem, groups_mem_kb, pkgs_dic)
        print "Finished: ##The output files is in: {}/".format(os.path.dirname(out_parse_table))
    else:
        original_dir = args.file[0]
        diff_dir = args.diff[0]
        meminfo_1 = "{}/{}".format(original_dir, meminfo_file)
        pm_list_1 = "{}/{}".format(original_dir, packages_file)
        meminfo_2 = "{}/{}".format(diff_dir, meminfo_file)
        pm_list_2 = "{}/{}".format(diff_dir, packages_file)

        system_mem_1, filtered_procs_mem_1 = get_file_meminfo(read_file(meminfo_1))
        system_mem_2, filtered_procs_mem_2 = get_file_meminfo(read_file(meminfo_2))
        pkgs_dic_1 = get_packages_dic(read_file(pm_list_1))
        pkgs_dic_2 = get_packages_dic(read_file(pm_list_2))
        groups_mem_kb_1, groups_mem_1 = get_procs_attr_group(filtered_procs_mem_1, pkgs_dic_1)
        groups_mem_kb_2, groups_mem_2 = get_procs_attr_group(filtered_procs_mem_2, pkgs_dic_2)
        print_diff_table(system_mem_1, system_mem_2, pkgs_dic_1, pkgs_dic_2,
                         groups_mem_kb_1, groups_mem_kb_2, groups_mem_1, groups_mem_2)

        # print "************* system meminfo*********************"
        # print system_mem_1
        # print "\n**************************************************"
        # print system_mem_2
        # print "\n\n************* process meminfo ********************"
        # print "filtered_procs_mem_1:"
        # for k, v in sorted(filtered_procs_mem_1.items(), key=lambda x: (-int(x[1]), x[0])):
        #     print (' {1}:  {0}'.format(k, v))
        #
        # print "\n**************************************************"
        # print "filtered_procs_mem_2:"
        # for k, v in sorted(filtered_procs_mem_2.items(), key=lambda x: (-int(x[1]), x[0])):
        #     print (' {1}:  {0}'.format(k, v))
        #
        # print "\n\n***************** packages ***********************"
        # print "pkgs_dic_1:"
        # for key in pkgs_dic_1:
        #     print key, ":"
        #     for pkg in sorted(pkgs_dic_1[key]):
        #         print "** ", pkg
        #     print 'len=', len(pkgs_dic_1[key])
        #
        # print "\n**************************************************"
        # print "pkgs_dic_2:"
        # for key in pkgs_dic_2:
        #     print key, ":"
        #     for pkg in sorted(pkgs_dic_2[key]):
        #         print "** ", pkg
        #     print 'len=', len(pkgs_dic_2[key])
        #
        # print "\n\n************** group meminfo kb ******************"
        # print groups_mem_kb_1
        # print "\n**************************************************"
        # print groups_mem_kb_2
        # print "\n\n**************group meminfo ***********************"
        # print "groups_mem_1:"
        # for key in groups_mem_1:
        #     print key, ":"
        #     for pkg in sorted(groups_mem_1[key]):
        #         print "** ", pkg
        #     print 'len=', len(groups_mem_1[key])
        # print "\n**************************************************"
        # print "groups_mem_2:"
        # for key in groups_mem_2:
        #     print key, ":"
        #     for pkg in sorted(groups_mem_2[key]):
        #         print "** ", pkg
        #     print 'len=', len(groups_mem_2[key])


def silent_remove(filename):
    try:
        if os.path.isdir(filename):
            shutil.rmtree(filename)
        else:
            os.remove(filename)
    except OSError as exc:  # this would be "except OSError, e:" before Python 2.6
        if exc.errno != errno.ENOENT:  # errno.ENOENT = no such file or directory
            raise  # re-raise exception if a different error occured


def read_file(filename):
    with open(filename, 'r') as fp:
        return fp.read()


def write_to_file(filename, file_meminfo):
    # automatically creating directories with file output
    if not os.path.exists(os.path.dirname(filename)):
        try:
            os.makedirs(os.path.dirname(filename))
        except OSError as exc:  # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise
    try:
        with open(filename, 'w') as fp:
            fp.write(file_meminfo)
    except IOError:
        print 'Oops! write file error.'
        silent_remove(filename)


def get_procs_attr_group(filtered_procs_mem, pkgs_dic):
    gms_kb = platform_kb = third_kb = sys_kb = 0
    groups_mem_kb = {}
    groups_mem = {}
    gms_procs_mem = []
    platform_procs_mem = []
    third_procs_mem = []
    sys_procs_mem = []

    for proc in filtered_procs_mem:
        if any(item in proc for item in GOOGLE_PROC_WHITE_LIST):
            gms_kb += int(filtered_procs_mem[proc])
            gms_procs_mem.append(proc)
        elif any(plat_pkg in proc for plat_pkg in PLATFORM_PKGS):
            platform_kb += int(filtered_procs_mem[proc])
            platform_procs_mem.append(proc)
        else:
            if any(pkg in proc for pkg in pkgs_dic['system'] if pkg != 'android' and pkg != 'android.overlay')\
                    or any(proc == item for item in SYSTEM_PROC_WHITE_LIST):
                sys_kb += int(filtered_procs_mem[proc])
                sys_procs_mem.append(proc)
            else:
                third_kb += int(filtered_procs_mem[proc])
                third_procs_mem.append(proc)

    groups_mem_kb.update({'GMS': gms_kb})
    groups_mem_kb.update({'Third': third_kb})
    groups_mem_kb.update({'System apps': sys_kb})
    if int(platform_flag):
        groups_mem_kb.update({'MTK': platform_kb})
    else:
        groups_mem_kb.update({'Qcom': platform_kb})

    groups_mem.update({'GMS': gms_procs_mem})
    groups_mem.update({'Third': third_procs_mem})
    groups_mem.update({'System apps': sys_procs_mem})
    if int(platform_flag):
        groups_mem.update({'MTK': platform_procs_mem})
    else:
        groups_mem.update({'Qcom': platform_procs_mem})
    return groups_mem_kb, groups_mem


def get_packages_dic(installed_pkgs_str):
    pkgs_dic = {}
    third_pkgs = []
    installed_pkgs = []
    for pkg in installed_pkgs_str.split('\n'):
        pkg_mat = re.search(r'package:(/\w+/[\w\-]+)/[^=]+=(.+)', pkg)
        if pkg_mat is not None:
            installed_pkgs.append(pkg_mat.group(2).strip())
            if "/data/app" == pkg_mat.group(1) and not any(item in pkg for item in GOOGLE_PROC_WHITE_LIST):
                third_pkgs.append(pkg_mat.group(2).strip())
    pkgs_dic.update({'third': third_pkgs})
    pkgs_dic.update({'installed': installed_pkgs})

    gms_pkgs = []
    platform_pkgs = []
    system_pkgs = []
    for pkg in installed_pkgs:
        if any(item in pkg for item in GOOGLE_PROC_WHITE_LIST):
            gms_pkgs.append(pkg)
        elif any(plat_pkg in pkg for plat_pkg in PLATFORM_PKGS):
            platform_pkgs.append(pkg)
        elif pkg not in third_pkgs:
            system_pkgs.append(pkg)
    pkgs_dic.update({'gms': gms_pkgs})
    pkgs_dic.update({'system': system_pkgs})
    if int(platform_flag):
        pkgs_dic.update({'mtk': platform_pkgs})
    else:
        pkgs_dic.update({'qcom': platform_pkgs})
    return pkgs_dic


def get_file_meminfo(file_meminfo_str):
    pss_mem_flag = False
    native_mem_flag = False
    system_mem = {}
    procs_mem = {}
    native_mem = set()

    mem_item_regx = re.compile('(\d+) kB: ([\w\.:_\-]*) \(pid.*')
    for line in file_meminfo_str.split('\n'):
        if "Total PSS by process:" in line:
            pss_mem_flag = True
        elif "Total PSS by OOM adjustment:" in line:
            pss_mem_flag = False
            native_mem_flag = True
        elif pss_mem_flag:
            proc = mem_item_regx.search(line)
            if proc is not None:
                procs_mem.update({proc.group(2): proc.group(1)})
        elif native_mem_flag:
            # Check whether the string is empty
            if not system_mem.get('Native', ""):
                native = re.search(r'(\d+) kB: Native', line)
                if native is not None:
                    system_mem.update({'Native': native.group(1)})
            elif re.search(r'\d+ kB: \w+?\s*[\r\n]+', line) is not None:
                native_mem_flag = False
            elif native_mem_flag:
                na = mem_item_regx.search(line)
                if na is not None:
                    native_mem.add(na.group(2))
        elif not system_mem.get('Total RAM', ""):
            total = re.search(r'Total RAM:\s+(\d+) kB', line)
            if total is not None:
                system_mem.update({'Total RAM': total.group(1)})
        elif not system_mem.get('Free RAM', ""):
            free = re.search(r'Free RAM:\s+(\d+) kB', line)
            if free is not None:
                system_mem.update({'Free RAM': free.group(1)})
        elif not system_mem.get('Used RAM', ""):
            used = re.search(r'Used RAM:\s+(\d+) kB\s+\((\d+) used pss\s+\+\s+(\d+) kernel', line)
            if used is not None:
                system_mem.update({'Used RAM': used.group(1)})
                system_mem.update({'Used Pss': used.group(2)})
                system_mem.update({'Kernel': used.group(3)})

    filtered_procs_mem = filter_out_native_processes(procs_mem, native_mem)
    return system_mem, filtered_procs_mem


def filter_out_native_processes(procs_mem, native_mem):
    filtered_procs_mem = {}
    for key in set(procs_mem.keys()) - native_mem:
        filtered_procs_mem.update({key: procs_mem[key]})
    return filtered_procs_mem


def print_mem_table(outfile, system_mem, groups_mem_kb, pkgs_dic):
    row_format = "{:<20}{}\n"
    if os.path.isfile(outfile):
        os.remove(outfile)
    with open(outfile, 'w') as fp:
        # write system memory info
        fp.write(row_format.format("*Memory*", "*Val(MB)*"))
        fp.write(row_format.format('Total RAM', kb2mb(system_mem['Total RAM'])))
        fp.write(row_format.format('Free RAM', kb2mb(system_mem['Free RAM'])))
        fp.write(row_format.format('Kernel', kb2mb(system_mem['Kernel'])))
        fp.write(row_format.format('Native', kb2mb(system_mem['Native'])))
        fp.write("\n")
        # write module memory info
        fp.write(row_format.format("*Module*", "*Val(MB)*"))
        fp.write(row_format.format('GMS', kb2mb(groups_mem_kb['GMS'])))
        fp.write(row_format.format('Third', kb2mb(groups_mem_kb['Third'])))
        fp.write(row_format.format('System apps', kb2mb(groups_mem_kb['System apps'])))
        if int(platform_flag):
            fp.write(row_format.format('MTK', kb2mb(groups_mem_kb['MTK'])))
        else:
            fp.write(row_format.format('Qcom', kb2mb(groups_mem_kb['Qcom'])))
        fp.write("\n")
        # write packages info
        fp.write(row_format.format("*Packages*", "*Count*"))
        fp.write(row_format.format('installed', len(pkgs_dic['installed'])))
        fp.write(row_format.format('gms', len(pkgs_dic['gms'])))
        fp.write(row_format.format('third', len(pkgs_dic['third'])))
        fp.write(row_format.format('system', len(pkgs_dic['system'])))
        if int(platform_flag):
            fp.write(row_format.format('mtk', len(pkgs_dic['mtk'])))
        else:
            fp.write(row_format.format('qcom', len(pkgs_dic['qcom'])))


def kb2mb(num):
    return int(num) // 1000


def styled_title_cell(ws, data):
    for c in data:
        c = ws.cell(column=data.index(c) + 1, row=1, value=c)
        c.fill = PatternFill(start_color='ff268bd2', end_color='ff268bd2', fill_type='solid')
        c.font = Font(name='Console', size=14, bold=True)
        c.alignment = Alignment(horizontal='left', vertical='center')
        yield c


def styled_subtitle(val, c):
    if val in ['Memory', 'Module', 'Packages']:
        c.font = Font(name='Courrier', size=12, bold=True)
    else:
        c.font = Font(name='Courrier', size=12)
    c.fill = PatternFill(patternType='solid', fgColor=Color(rgb='00C5D9F1'))
    c.alignment = Alignment(horizontal='left', vertical='center')


def styled_cell(c):
    c.font = Font(name='Courrier', size=12)
    c.alignment = Alignment(horizontal='left', vertical='top')


def render_ws(ws):
    # adjust width of the column
    for i in range(ws.max_column):
        ws.column_dimensions[get_column_letter(i + 1)].width = 20

    # adjust height of the row
    for i in range(ws.max_row):
        # Get value of specific cells with openpyxl
        val = ws["{}{}".format(get_column_letter(5), i + 1)].value
        if i == 0:
            ws.row_dimensions[i + 1].height = 23
        elif val is not None:
            if len(val.split('\n')) <= 10:
                ws.row_dimensions[i + 1].height = 18 * len(val.split('\n'))
            else:
                ws.row_dimensions[i + 1].height = 18 * 10
        else:
            ws.row_dimensions[i + 1].height = 18

    # set border, used "iter_rows"
    for row in ws.iter_rows("A1:E17"):
        for i in range(ws.max_column):
            row[i].border = Border(left=Side(border_style='thin'),
                          right=Side(border_style='thin'),
                          top=Side(border_style='thin'),
                          bottom=Side(border_style='thin'))


def print_diff_table(system_mem_1, system_mem_2, pkgs_dic_1, pkgs_dic_2,
                     groups_mem_kb_1, groups_mem_kb_2, groups_mem_1, groups_mem_2):
    ws_tile = ['Items', original_dir, diff_dir, 'Gap', 'Detail']
    total_memory_items = ['Total RAM', 'Free RAM', 'Kernel', 'Native']
    if int(platform_flag):
        module_memory_items = ['GMS', 'MTK', 'Third', 'System apps']
        package_items = ['installed', 'gms', 'mtk', 'third', 'system']
    else:
        module_memory_items = ['GMS', 'Qcom', 'Third', 'System apps']
        package_items = ['installed', 'gms', 'qcom', 'third', 'system']
    wb = Workbook()
    ws = wb.active
    ws.title = "Memory Diff Table"
    # Coloring a tab in openpyxl
    ws.sheet_properties.tabColor = 'FFFF9900'  # Orange
    # add title at the top of the table
    ws.append(styled_title_cell(ws, ws_tile))

    if int(platform_flag):
        title = MTK_TITLE
    else:
        title = QCOM_TITLE

    for col in title:
        styled_subtitle(col, ws.cell(column=1, row=title.index(col) + 2, value=col))
        if col in total_memory_items:
            styled_cell(ws.cell(column=2, row=title.index(col) + 2, value="{}MB".format(kb2mb(system_mem_1[col]))))
            styled_cell(ws.cell(column=3, row=title.index(col) + 2, value="{}MB".format(kb2mb(system_mem_2[col]))))
            styled_cell(ws.cell(column=4, row=title.index(col) + 2,
                                value="{}MB".format(kb2mb(system_mem_1[col]) - kb2mb(system_mem_2[col]))))
        if col in module_memory_items:
            styled_cell(ws.cell(column=2, row=title.index(col) + 2, value="{}MB".format(kb2mb(groups_mem_kb_1[col]))))
            styled_cell(ws.cell(column=3, row=title.index(col) + 2, value="{}MB".format(kb2mb(groups_mem_kb_2[col]))))
            styled_cell(ws.cell(column=4, row=title.index(col) + 2,
                                value="{}MB".format(kb2mb(groups_mem_kb_1[col]) - kb2mb(groups_mem_kb_2[col]))))
            c = ws.cell(column=5, row=title.index(col) + 2)
            # c.style.alignment.wrap_text = True
            c.value = "ADDED PROCESSES:\n{}\n\nREMOVED PROCESSES:\n{}".format(
                '\n'.join(set(groups_mem_1[col]) - set(groups_mem_2[col])),
                '\n'.join(set(groups_mem_2[col]) - set(groups_mem_1[col])))
            styled_cell(c)
        if col in package_items:
            styled_cell(ws.cell(column=2, row=title.index(col) + 2, value=len(pkgs_dic_1[col])))
            styled_cell(ws.cell(column=3, row=title.index(col) + 2, value=len(pkgs_dic_2[col])))
            styled_cell(ws.cell(column=4, row=title.index(col) + 2,
                                value=len(pkgs_dic_1[col]) - len(pkgs_dic_2[col])))
            c = ws.cell(column=5, row=title.index(col) + 2)
            # c.style.alignment.wrap_text = True
            diff_pkgs = [['\nADDED PACKAGES:'], set(pkgs_dic_1[col]) - set(pkgs_dic_2[col]),
                         ['\nREMOVED PACKAGES:'], set(pkgs_dic_2[col]) - set(pkgs_dic_1[col])]
            # extract nested lists
            merged = list(itertools.chain(*diff_pkgs))
            if len(merged) != 0:
                c.value = '\n'.join(merged)
                styled_cell(c)

    render_ws(ws)
    wb.save(filename=out_parse_xlsx)
    print "Finished: save xlsx at", out_parse_xlsx


def filter_out_items(list1, list2):
    set(list1) - set(list2)


if __name__ == '__main__':
    # Catching ANY exception, and remove output files if any.
    try:
        main()
    except Exception as e:
        silent_remove(os.path.dirname(out_parse_table))
        silent_remove(out_parse_xlsx)
        raise
