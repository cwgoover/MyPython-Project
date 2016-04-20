#!/usr/bin/python
import argparse
import itertools
import ntpath
import os
import re
from subprocess import check_output

from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.compat import range
from openpyxl.styles import Color, PatternFill, Font, Alignment, Border, Side

# Add the PATH environment variable to your Run Configuration
# (Run->Edit Configurations) like this: PATH=/local/bin:$PATH
#
# cmd_meminfo = "/local/bin/adb shell dumpsys meminfo"

# The shell argument (which defaults to False) specifies whether to use the shell as
# the program to execute. If shell is True, it is recommended to pass args as a string
# rather than as a sequence.
#
# cmd_meminfo = ["adb", "shell", "dumpsys", "meminfo", "|", "tee", "meminfo.txt"]

# In practice it means you can pass the arguments as a string, instead of a list,
# as check_output() would normally expect.
#
# p = Popen(cmd_meminfo, stdout=PIPE, shell=True)
# output, err = p.communicate()

CMD_PACKAGES = "adb shell pm list packages -f | tee pkgs.txt"
CMD_MEMINFO = "adb shell dumpsys meminfo | tee meminfo.txt"

TITLE = ['Memory', 'Total RAM', 'Free RAM', 'Kernel', 'Native',
         'Module', 'GMS', 'Qcom', 'Third', 'System apps',
         'Packages', 'installed', 'gms', 'qcom', 'third', 'system']

meminfo_1 = ""
meminfo_2 = ""
outfile = ""


def main():
    global meminfo_1, meminfo_2, outfile
    # TODO: need usage()

    # compare: xx.py -f meminfo pm -d meminfo pm -o output
    # analyse: xx.py -o table (need default file name)
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', nargs='*', help='need meminfo & pm info two files')
    parser.add_argument('-d', '--diff', nargs='*', help='need meminfo & pm info two files')
    parser.add_argument('-o', '--output', nargs='?', default='parse_result.txt',
                        help='output file with table format')
    args = parser.parse_args()
    outfile = str(args.output)

    if args.file is None:
        # Run command with arguments and return its output as a byte string.
        file_meminfo = check_output(CMD_MEMINFO, shell=True)
        system_mem, filtered_procs_mem = get_file_meminfo(file_meminfo)

        installed_pkgs_str = check_output(CMD_PACKAGES, shell=True)
        pkgs_dic = get_packages_dic(installed_pkgs_str)

        groups_mem_kb, groups_mem = get_procs_attr_group(filtered_procs_mem, pkgs_dic)
        print_mem_table(system_mem, groups_mem_kb, pkgs_dic)
    else:
        meminfo_1 = args.file[0]
        pm_list_1 = args.file[1]
        meminfo_2 = args.diff[0]
        pm_list_2 = args.diff[1]
        # change output file to .xlsx
        base = ntpath.basename(outfile)
        outfile = "{}.xlsx".format(ntpath.splitext(base)[0])

        system_mem_1, filtered_procs_mem_1 = get_file_meminfo(read_file(meminfo_1))
        system_mem_2, filtered_procs_mem_2 = get_file_meminfo(read_file(meminfo_2))
        pkgs_dic_1 = get_packages_dic(read_file(pm_list_1))
        pkgs_dic_2 = get_packages_dic(read_file(pm_list_2))
        groups_mem_kb_1, groups_mem_1 = get_procs_attr_group(filtered_procs_mem_1, pkgs_dic_1)
        groups_mem_kb_2, groups_mem_2 = get_procs_attr_group(filtered_procs_mem_2, pkgs_dic_2)
        print_diff_table(system_mem_1, system_mem_2, pkgs_dic_1, pkgs_dic_2,
                         groups_mem_kb_1, groups_mem_kb_2, groups_mem_1, groups_mem_2)
        # print "************* meminfo*********************"
        # print system_mem_1
        # print "******************************************"
        # print system_mem_2
        # print "******************************************"
        # for k, v in sorted(filtered_procs_mem_1.items(), key=lambda x: (-int(x[1]), x[0])):
        #     print (' {1}:  {0}'.format(k, v))
        #
        # print "******************************************"
        # for k, v in sorted(filtered_procs_mem_2.items(), key=lambda x: (-int(x[1]), x[0])):
        #     print (' {1}:  {0}'.format(k, v))
        #
        # print "************* packages ***********************"
        # for key in pkgs_dic_1:
        #     print key, ":"
        #     for pkg in sorted(pkgs_dic_1[key]):
        #         print "** ", pkg
        #     print 'len=', len(pkgs_dic_1[key])
        #
        # print "******************************************"
        # for key in pkgs_dic_2:
        #     print key, ":"
        #     for pkg in sorted(pkgs_dic_2[key]):
        #         print "** ", pkg
        #     print 'len=', len(pkgs_dic_2[key])
        #
        # print "******************************************"
        # print groups_mem_kb_1
        # print "******************************************"
        # print groups_mem_kb_2
        # print "******************************************"
        # print groups_mem_1
        # print "******************************************"
        # print groups_mem_2


def read_file(filename):
    with open(filename, 'r') as fp:
        return fp.read()


def get_procs_attr_group(filtered_procs_mem, pkgs_dic):
    gms_kb = qcom_kb = third_kb = sys_kb = 0
    groups_mem_kb = {}
    groups_mem = {}
    gms_procs_mem = []
    qcom_procs_mem = []
    third_procs_mem = []
    sys_procs_mem = []
    for proc in filtered_procs_mem:
        if "com.google" in proc or "chrome" in proc:
            gms_kb += int(filtered_procs_mem[proc])
            gms_procs_mem.append(proc)
        elif "com.qti" in proc or "com.qualcomm" in proc:
            qcom_kb += int(filtered_procs_mem[proc])
            qcom_procs_mem.append(proc)
        else:
            # prefix_third = re.search(r'\w+\.\w+', proc)
            # Check if substring is in a list of strings
            # http://stackoverflow.com/a/16380333/4710864
            # if prefix_third is not None and any(prefix_third.group() in pkg for pkg in third_pkgs):
            #
            # If any item of list starts with string?
            #
            # if there is only one "com.tct" in the third packages, resulted in all the com.tct pkgs in the third
            # if "com.jrdcom" not in proc and prefix_third is not None and any(
            #         pkg.startswith(prefix_third.group()) for pkg in pkgs_dic['third']):
            if any(pkg in proc for pkg in pkgs_dic['system']) or proc == 'system':
                sys_kb += int(filtered_procs_mem[proc])
                sys_procs_mem.append(proc)
            else:
                third_kb += int(filtered_procs_mem[proc])
                third_procs_mem.append(proc)

    groups_mem_kb.update({'GMS': gms_kb})
    groups_mem_kb.update({'Qcom': qcom_kb})
    groups_mem_kb.update({'Third': third_kb})
    groups_mem_kb.update({'System apps': sys_kb})

    groups_mem.update({'GMS': gms_procs_mem})
    groups_mem.update({'Qcom': qcom_procs_mem})
    groups_mem.update({'Third': third_procs_mem})
    groups_mem.update({'System apps': sys_procs_mem})
    return groups_mem_kb, groups_mem


def get_packages_dic(installed_pkgs_str):
    pkgs_dic = {}
    third_pkgs = []
    installed_pkgs = []
    for pkg in installed_pkgs_str.split('\n'):
        pkg_mat = re.search(r'package:(/\w+/[\w\-]+)/[^=]+=(.+)', pkg)
        if pkg_mat is not None:
            installed_pkgs.append(pkg_mat.group(2).strip())
            if "/data/app" == pkg_mat.group(1):
                third_pkgs.append(pkg_mat.group(2).strip())
    pkgs_dic.update({'third': third_pkgs})
    pkgs_dic.update({'installed': installed_pkgs})

    gms_pkgs = []
    qcom_pkgs = []
    system_pkgs = []
    for pkg in installed_pkgs:
        if "com.google" in pkg or "chrome" in pkg:
            gms_pkgs.append(pkg)
        elif "com.qti" in pkg or "com.qualcomm" in pkg:
            qcom_pkgs.append(pkg)
        elif pkg not in third_pkgs:
            system_pkgs.append(pkg)
    pkgs_dic.update({'gms': gms_pkgs})
    pkgs_dic.update({'qcom': qcom_pkgs})
    pkgs_dic.update({'system': system_pkgs})
    return pkgs_dic


def get_file_meminfo(file_meminfo_str):
    pss_mem_flag = False
    native_mem_flag = False
    system_mem = {}
    procs_mem = {}
    # How can I add items to an empty set in python
    # http://stackoverflow.com/a/17511281/4710864
    native_mem = set()

    mem_item_regx = re.compile('(\d+) kB: ([\w\.:_\-]*) \(pid.*')
    for line in file_meminfo_str.split('\n'):
        # A string contains substring method
        if "Total PSS by process:" in line:
            pss_mem_flag = True
        elif "Total PSS by OOM adjustment:" in line:
            pss_mem_flag = False
            native_mem_flag = True
        elif pss_mem_flag:
            proc = mem_item_regx.search(line)
            if proc is not None:
                procs_mem.update({proc.group(2): proc.group(1)})
        elif native_mem_flag:
            # Check whether the string is empty
            if not system_mem.get('Native', ""):
                native = re.search(r'(\d+) kB: Native', line)
                if native is not None:
                    system_mem.update({'Native': native.group(1)})
            # elif re.search(r'\d+ kB: Persistent', line) is not None:
            elif re.search(r'\d+ kB: \w+?\s*[\r\n]+', line) is not None:
                native_mem_flag = False
            elif native_mem_flag:
                na = mem_item_regx.search(line)
                if na is not None:
                    native_mem.add(na.group(2))
        elif not system_mem.get('Total RAM', ""):
            total = re.search(r'Total RAM:\s+(\d+) kB', line)
            if total is not None:
                system_mem.update({'Total RAM': total.group(1)})
        elif not system_mem.get('Free RAM', ""):
            free = re.search(r'Free RAM:\s+(\d+) kB', line)
            if free is not None:
                system_mem.update({'Free RAM': free.group(1)})
        elif not system_mem.get('Used RAM', ""):
            used = re.search(r'Used RAM:\s+(\d+) kB\s+\((\d+) used pss\s+\+\s+(\d+) kernel\)', line)
            if used is not None:
                system_mem.update({'Used RAM': used.group(1)})
                system_mem.update({'Used Pss': used.group(2)})
                system_mem.update({'Kernel': used.group(3)})

    if __debug__:
        pss_kb = 0
        # # Python Sorted: Sorting a dictionary by value (DESC) then by key (ASC)
        # # http://stackoverflow.com/a/15371752/4710864
        # for k, v in sorted(procs_mem.items(), key=lambda x: (-int(x[1]), x[0])):
        #     print (' {1}:  {0}'.format(k, v))
        #     pss_kb += int(v)
        # print 'procs_mem size=', len(procs_mem), ', pss_kb=', pss_kb
        # print '\n************\n'
        # print 'system_mem=', system_mem
        # print '\n************\n'

    filtered_procs_mem = filter_out_native_processes(procs_mem, native_mem)
    return system_mem, filtered_procs_mem


def filter_out_native_processes(procs_mem, native_mem):
    filtered_procs_mem = {}
    for key in set(procs_mem.keys()) - native_mem:
        filtered_procs_mem.update({key: procs_mem[key]})
    return filtered_procs_mem


def print_mem_table(system_mem, groups_mem_kb, pkgs_dic):
    row_format = "{:<20}{}\n"
    if os.path.isfile(outfile):
        os.remove(outfile)
    with open(outfile, 'w') as fp:
        # write system memory info
        fp.write(row_format.format("*Memory*", "*Val(MB)*"))
        fp.write(row_format.format('Total RAM', kb2mb(system_mem['Total RAM'])))
        fp.write(row_format.format('Free RAM', kb2mb(system_mem['Free RAM'])))
        fp.write(row_format.format('Kernel', kb2mb(system_mem['Kernel'])))
        fp.write(row_format.format('Native', kb2mb(system_mem['Native'])))
        fp.write("\n")
        # write module memory info
        fp.write(row_format.format("*Module*", "*Val(MB)*"))
        fp.write(row_format.format('GMS', kb2mb(groups_mem_kb['GMS'])))
        fp.write(row_format.format('Qcom', kb2mb(groups_mem_kb['Qcom'])))
        fp.write(row_format.format('Third', kb2mb(groups_mem_kb['Third'])))
        fp.write(row_format.format('System apps', kb2mb(groups_mem_kb['System apps'])))
        fp.write("\n")
        # write packages info
        fp.write(row_format.format("*Packages*", "*Count*"))
        fp.write(row_format.format('installed', len(pkgs_dic['installed'])))
        fp.write(row_format.format('gms', len(pkgs_dic['gms'])))
        fp.write(row_format.format('qcom', len(pkgs_dic['qcom'])))
        fp.write(row_format.format('third', len(pkgs_dic['third'])))
        fp.write(row_format.format('system', len(pkgs_dic['system'])))


def kb2mb(num):
    return int(num) // 1000


# border_style = Style(font=Font(name='Console', size=10, bold=False,
#                                color=Color(openpyxl.styles.colors.BLACK)),
#                      fill=PatternFill(patternType='solid', fgColor=Color(rgb='00C5D9F1')),
#                      border=Border(bottom=Side(border_style='medium', color=Color(rgb='FF000000'))))
def styled_title_cell(ws, data):
    for c in data:
        c = ws.cell(column=data.index(c) + 1, row=1, value=c)
        c.fill = PatternFill(start_color='ff268bd2', end_color='ff268bd2', fill_type='solid')
        c.font = Font(name='Console', size=14, bold=True)
        c.alignment = Alignment(horizontal='left', vertical='center')
        yield c


def styled_subtitle(val, c):
    if val in ['Memory', 'Module', 'Packages']:
        c.font = Font(name='Courrier', size=12, bold=True)
    else:
        c.font = Font(name='Courrier', size=12)
    c.fill = PatternFill(patternType='solid', fgColor=Color(rgb='00C5D9F1'))
    c.alignment = Alignment(horizontal='left', vertical='center')


def styled_cell(c):
    c.font = Font(name='Courrier', size=12)
    c.alignment = Alignment(horizontal='left', vertical='top')


def render_ws(ws):
    # adjust width of the column
    for i in range(ws.max_column):
        ws.column_dimensions[get_column_letter(i + 1)].width = 20

    # adjust height of the row
    # http://stackoverflow.com/q/32855656/4710864
    for i in range(ws.max_row):
        # Get value of specific cells with openpyxl
        # http://stackoverflow.com/a/29157762/4710864
        val = ws["{}{}".format(get_column_letter(5), i + 1)].value
        if i == 0:
            ws.row_dimensions[i + 1].height = 23
        elif val is not None:
            if len(val.split('\n')) <= 10:
                ws.row_dimensions[i + 1].height = 18 * len(val.split('\n'))
            else:
                ws.row_dimensions[i + 1].height = 18 * 10
        else:
            ws.row_dimensions[i + 1].height = 18

    # set border, used "iter_rows"
    for row in ws.iter_rows("A1:E17"):
        for i in range(ws.max_column):
            row[i].border = Border(left=Side(border_style='thin'),
                          right=Side(border_style='thin'),
                          top=Side(border_style='thin'),
                          bottom=Side(border_style='thin'))


def print_diff_table(system_mem_1, system_mem_2, pkgs_dic_1, pkgs_dic_2,
                     groups_mem_kb_1, groups_mem_kb_2, groups_mem_1, groups_mem_2):
    ws_tile = ['Items', ntpath.basename(meminfo_1), ntpath.basename(meminfo_2), 'Gap', 'Detail']
    total_memory_items = ['Total RAM', 'Free RAM', 'Kernel', 'Native']
    module_memory_items = ['GMS', 'Qcom', 'Third', 'System apps']
    package_items = ['installed', 'gms', 'qcom', 'third', 'system']

    wb = Workbook()
    ws = wb.active
    ws.title = "Memory Diff Table"
    # Coloring a tab in openpyxl
    ws.sheet_properties.tabColor = 'FFFF9900'  # Orange
    # add title at the top of the table
    ws.append(styled_title_cell(ws, ws_tile))

    for col in TITLE:
        styled_subtitle(col, ws.cell(column=1, row=TITLE.index(col) + 2, value=col))
        if col in total_memory_items:
            styled_cell(ws.cell(column=2, row=TITLE.index(col) + 2, value="{}MB".format(kb2mb(system_mem_1[col]))))
            styled_cell(ws.cell(column=3, row=TITLE.index(col) + 2, value="{}MB".format(kb2mb(system_mem_2[col]))))
            styled_cell(ws.cell(column=4, row=TITLE.index(col) + 2,
                                value="{}MB".format(kb2mb(system_mem_1[col]) - kb2mb(system_mem_2[col]))))
        if col in module_memory_items:
            styled_cell(ws.cell(column=2, row=TITLE.index(col) + 2, value="{}MB".format(kb2mb(groups_mem_kb_1[col]))))
            styled_cell(ws.cell(column=3, row=TITLE.index(col) + 2, value="{}MB".format(kb2mb(groups_mem_kb_2[col]))))
            styled_cell(ws.cell(column=4, row=TITLE.index(col) + 2,
                                value="{}MB".format(kb2mb(groups_mem_kb_1[col]) - kb2mb(groups_mem_kb_2[col]))))
            # Writing multi-line strings into cells using openpyxl
            # http://stackoverflow.com/questions/15370432/writing-multi-line-strings-into-cells-using-openpyxl
            c = ws.cell(column=5, row=TITLE.index(col) + 2)
            c.style.alignment.wrap_text = True
            c.value = "ADDED PROCESSES:\n{}".format('\n'.join(set(groups_mem_1[col]) - set(groups_mem_2[col])))
            styled_cell(c)
        if col in package_items:
            styled_cell(ws.cell(column=2, row=TITLE.index(col) + 2, value=len(pkgs_dic_1[col])))
            styled_cell(ws.cell(column=3, row=TITLE.index(col) + 2, value=len(pkgs_dic_2[col])))
            styled_cell(ws.cell(column=4, row=TITLE.index(col) + 2,
                                value=len(pkgs_dic_1[col]) - len(pkgs_dic_2[col])))
            c = ws.cell(column=5, row=TITLE.index(col) + 2)
            c.style.alignment.wrap_text = True
            diff_pkgs = [['\nADDED PACKAGES:'], set(pkgs_dic_1[col]) - set(pkgs_dic_2[col]),
                         ['\nREMOVED PACKAGES:'], set(pkgs_dic_2[col]) - set(pkgs_dic_1[col])]
            # how to extract nested lists?
            # http://stackoverflow.com/a/953097/4710864
            merged = list(itertools.chain(*diff_pkgs))
            if len(merged) != 0:
                c.value = '\n'.join(merged)
                styled_cell(c)

    render_ws(ws)
    wb.save(filename=outfile)


def filter_out_items(list1, list2):
    set(list1) - set(list2)


if __name__ == '__main__':
    main()
